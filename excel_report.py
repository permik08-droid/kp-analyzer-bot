import re
from procurement_risks import analyze_procurement_risks
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment


def get_position_key(name: str) -> str:
    text = str(name).upper()

    text = text.replace("×", "X")
    text = text.replace("Х", "X")
    text = text.replace("-", " ")
    text = text.replace("CKB", "СКВ")
    text = text.replace("CK", "СК")
    text = re.sub(r"[.,;:(){}\[\]\"']", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    patterns = [
        r"BT\d+\s+СКВ\d+\s+\d+",
        r"BT\d+\s+CKB\d+\s+\d+",
        r"СКВ\d+\s+TWN\d+\s+\d+",
        r"CKB\d+\s+TWN\d+\s+\d+",
        r"RDH\s+D\d+\s+\d+\s+\d+L",
        r"PS\s+BT\d+\s+\d+\s+HO",
        r"[A-ZА-Я]{2,}\d+\s+\d+"
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0).strip()

    noise_words = [
        "КУПИТЬ",
        "ПОСТАВКА",
        "ТОВАР",
        "ИЗДЕЛИЕ",
        "МАТЕРИАЛ",
        "КОМПЛЕКТ",
        "ШТ",
        "ШТ.",
        "ЕД",
        "ЕД.",
        "РУБ",
        "ТГ",
        "ТЕНГЕ"
    ]

    words = text.split()
    words = [word for word in words if word not in noise_words]

    return " ".join(words).strip()


def extract_article(name):
    text = str(name).upper()
    text = text.replace("×", "X")
    text = text.replace("Х", "X")
    text = text.replace("-", " ")
    text = re.sub(r"[.,;:(){}\[\]\"']", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    patterns = [
        r"BT\d+\s+СКВ\d+\s+\d+",
        r"BT\d+\s+CKB\d+\s+\d+",
        r"СКВ\d+\s+TWN\d+\s+\d+",
        r"CKB\d+\s+TWN\d+\s+\d+",
        r"RDH\s+D\d+\s+\d+\s+\d+L",
        r"PS\s+BT\d+\s+\d+\s+HO",
        r"M\d+\s*X\s*\d+",
        r"\d+\s*X\s*\d+"
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(0).strip()

    return ""


def normalize_unit(unit):
    value = str(unit).strip().lower()
    value = value.replace(".", "")

    unit_map = {
        "шт": "шт",
        "штука": "шт",
        "штук": "шт",
        "ед": "шт",
        "единица": "шт",
        "единиц": "шт",
        "комплект": "компл",
        "комплекта": "компл",
        "компл": "компл",
        "м": "м",
        "метр": "м",
        "метра": "м",
        "пм": "м",
        "пог м": "м",
        "кг": "кг",
        "килограмм": "кг",
        "тонна": "т",
        "т": "т",
        "л": "л",
        "литр": "л"
    }

    return unit_map.get(value, value)


def clean_price(price):
    try:
        value = str(price)
        value = value.replace(" ", "")
        value = value.replace(",", ".")
        value = re.sub(r"[^0-9.]", "", value)
        if value == "":
            return None
        return float(value)
    except Exception:
        return None


def style_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(border_style="thin", color="999999")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_procurement_report(items: list, output_path: str):
    wb = Workbook()

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

    supplier_stats = {}

    # Лист 1 — Сравнение КП
    ws = wb.active
    ws.title = "Сравнение КП"

    ws.append([
        "№",
        "Поставщик",
        "Сумма",
        "НДС",
        "Доставка",
        "Срок поставки",
        "Условия оплаты",
        "Гарантия",
        "Срок действия КП",
        "Производитель",
        "Страна",
        "Комментарий"
    ])

    for index, item in enumerate(items, start=1):
        ws.append([
            index,
            item.get("supplier", "не указано"),
            item.get("total_amount", "не указано"),
            item.get("vat", "не указано"),
            item.get("delivery", "не указано"),
            item.get("delivery_time", "не указано"),
            item.get("payment_terms", "не указано"),
            item.get("warranty", "не указано"),
            item.get("valid_until", "не указано"),
            item.get("manufacturer", "не указано"),
            item.get("country", "не указано"),
            item.get("comment", "")
        ])

    style_sheet(ws, {
        "A": 6,
        "B": 30,
        "C": 18,
        "D": 15,
        "E": 35,
        "F": 25,
        "G": 28,
        "H": 18,
        "I": 22,
        "J": 25,
        "K": 18,
        "L": 45
    })

    # Лист 2 — Позиции КП
    ws_items = wb.create_sheet("Позиции КП")

    ws_items.append([
        "Файл",
        "Поставщик",
        "Ключ позиции",
        "Наименование",
        "Ед. изм.",
        "Кол-во",
        "Цена",
        "Сумма"
    ])

    for item in items:
        supplier = item.get("supplier", "не указано")
        file_name = item.get("file_name", "не указано")

        if supplier not in supplier_stats:
            supplier_stats[supplier] = {
                "wins": 0,
                "sum_wins": 0,
                "positions": 0
            }

        for position in item.get("items", []):
            name = position.get("name", "не указано")
            key_source = (
                position.get("match_key")
                or position.get("normalized_name")
                or name
            )
            key = get_position_key(key_source)

            supplier_stats[supplier]["positions"] += 1

            ws_items.append([
                file_name,
                supplier,
                key,
                name,
                position.get("unit", "не указано"),
                position.get("quantity", "не указано"),
                position.get("price", "не указано"),
                position.get("amount", "не указано")
            ])

    style_sheet(ws_items, {
        "A": 35,
        "B": 30,
        "C": 25,
        "D": 55,
        "E": 12,
        "F": 12,
        "G": 18,
        "H": 18
    })

    # Лист 3 — Конкурентная карта
    ws_map = wb.create_sheet("Конкурентная карта")

    suppliers = []
    position_data = {}

    for item in items:
        supplier = item.get("supplier", "не указано")

        if supplier not in suppliers:
            suppliers.append(supplier)

        for position in item.get("items", []):
            name = position.get("name", "не указано")
            key_source = (
                position.get("match_key")
                or position.get("normalized_name")
                or name
            )
            key = get_position_key(key_source)
            price = position.get("price", "не указано")

            if key not in position_data:
                position_data[key] = {
                    "name": name,
                    "prices": {}
                }

            position_data[key]["prices"][supplier] = price

    ws_map.append(
        ["Ключ позиции", "Наименование"]
        + suppliers
        + ["Победитель", "Мин. цена", "Макс. цена", "Экономия"]
    )

    total_saving = 0

    for key, data in position_data.items():
        row = [key, data["name"]]
        found_prices = []

        for supplier in suppliers:
            price = data["prices"].get(supplier, "")
            row.append(price)

            numeric_price = clean_price(price)
            if numeric_price is not None:
                found_prices.append((numeric_price, supplier))

        if found_prices:
            min_price, winner = min(found_prices, key=lambda x: x[0])
            max_price, _ = max(found_prices, key=lambda x: x[0])
            saving = max_price - min_price

            row.append(winner)
            row.append(min_price)
            row.append(max_price)
            row.append(saving)

            total_saving += saving

            if winner not in supplier_stats:
                supplier_stats[winner] = {
                    "wins": 0,
                    "sum_wins": 0,
                    "positions": 0
                }

            supplier_stats[winner]["wins"] += 1
            supplier_stats[winner]["sum_wins"] += min_price
        else:
            row.append("не определён")
            row.append("")
            row.append("")
            row.append("")

        ws_map.append(row)

    style_sheet(ws_map, {
        "A": 25,
        "B": 55,
        "C": 22,
        "D": 22,
        "E": 22,
        "F": 22,
        "G": 22,
        "H": 30,
        "I": 18,
        "J": 18,
        "K": 18
    })

    # Подсветка минимальных цен
    first_supplier_col = 3
    last_supplier_col = 2 + len(suppliers)

    for row_idx in range(2, ws_map.max_row + 1):
        prices = []

        for col_idx in range(first_supplier_col, last_supplier_col + 1):
            value = ws_map.cell(row=row_idx, column=col_idx).value
            numeric = clean_price(value)
            if numeric is not None:
                prices.append((numeric, col_idx))

        if prices:
            min_price, min_col = min(prices, key=lambda x: x[0])

            ws_map.cell(
                row=row_idx,
                column=min_col
            ).fill = green_fill

            avg_price = sum(p[0] for p in prices) / len(prices)

            if min_price < avg_price * 0.7:
                cell = ws_map.cell(
                    row=row_idx,
                    column=min_col
                )

                cell.fill = red_fill

                cell.comment = Comment(
                    "Аномально низкая цена. Проверьте комплектацию и условия поставки.",
                    "KP Bot"
                )
    # Лист 4 — Сопоставление позиций
    ws_matching = wb.create_sheet("Сопоставление позиций")

    ws_matching.append([
        "Ключ позиции",
        "Поставщик",
        "Исходное наименование",
        "match_key GPT",
        "normalized_name",
        "Ед. изм.",
        "Кол-во",
        "Цена",
        "Сумма"
    ])

    matching_groups = {}
    for item in items:
        supplier = item.get("supplier", "не указано")

        for position in item.get("items", []):
            name = position.get("name", "не указано")
            key_source = (
                position.get("match_key")
                or position.get("normalized_name")
                or name
            )
            key = get_position_key(key_source)
            if key not in matching_groups:
                matching_groups[key] = []

            matching_groups[key].append({
                "supplier": supplier,
                "name": name,
                "unit": position.get("unit", "не указано"),
                "quantity": position.get("quantity", "не указано"),
                "price": position.get("price", "не указано"),
                "amount": position.get("amount", "не указано"),
                "match_key": position.get("match_key", ""),
                "normalized_name": position.get("normalized_name", "")
            })
            ws_matching.append([
                key,
                supplier,
                name,
                position.get("match_key", ""),
                position.get("normalized_name", ""),
                position.get("unit", "не указано"),
                position.get("quantity", "не указано"),
                position.get("price", "не указано"),
                position.get("amount", "не указано")
            ])

    style_sheet(ws_matching, {
        "A": 25,
        "B": 30,
        "C": 60,
        "D": 35,
        "E": 35,
        "F": 12,
        "G": 12,
        "H": 18,
        "I": 18
    })
    # Лист 5 — Контроль сопоставления
    ws_match_control = wb.create_sheet("Контроль сопоставления")

    ws_match_control.append([
        "Ключ позиции",
        "Проблема",
        "Уровень",
        "Детали"
    ])

    for key, group in matching_groups.items():
        units = sorted(set(
            normalize_unit(position.get("unit", "не указано"))
            for position in group
            if str(position.get("unit", "не указано")).strip()
        ))

        if len(units) > 1:
            ws_match_control.append([
                key,
                "Разные единицы измерения",
                "Предупреждение",
                ", ".join(units)
            ])
        quantity_details = sorted(set(
            f"{position.get('supplier', 'не указано')}: {position.get('quantity', 'не указано')}"
            for position in group
            if str(position.get("quantity", "не указано")).strip()
        ))

        quantities = sorted(set(
            str(position.get("quantity", "не указано")).strip()
            for position in group
            if str(position.get("quantity", "не указано")).strip()
        ))

        if len(quantities) > 1:
            ws_match_control.append([
                key,
                "Разные количества",
                "Критично",
                " | ".join(quantity_details)
            ])
        price_details = []

        for position in group:
            numeric_price = clean_price(position.get("price", "не указано"))

            if numeric_price is not None:
                price_details.append(
                    (
                        numeric_price,
                        position.get("supplier", "не указано")
                    )
                )

        if len(price_details) > 1:
            min_price, min_supplier = min(price_details, key=lambda x: x[0])
            max_price, max_supplier = max(price_details, key=lambda x: x[0])

            if min_price > 0 and max_price / min_price > 3:
                ratio = round(max_price / min_price, 1)

                ws_match_control.append([
                    key,
                    "Большой разброс цен",
                    "Предупреждение",
                    f"{min_supplier}: {min_price} | {max_supplier}: {max_price} | x{ratio}"
                ])
        names = sorted(set(
            str(position.get("name", "не указано")).strip()
            for position in group
            if str(position.get("name", "не указано")).strip()
        ))

        if len(names) > 3:
            ws_match_control.append([
                key,
                "Много разных наименований в одной группе",
                "Предупреждение",
                " | ".join(names[:5])
            ])
        articles = sorted(set(
            extract_article(position.get("name", ""))
            for position in group
            if extract_article(position.get("name", ""))
        ))

        if len(articles) > 1:
            ws_match_control.append([
                key,
                "Разные артикулы в одной группе",
                "Критично",
                " | ".join(articles[:5])
            ])
    style_sheet(ws_match_control, {
        "A": 30,
        "B": 35,
        "C": 18,
        "D": 80
    })
    for row in ws_match_control.iter_rows(min_row=2):
        level = row[2].value

        if level == "Критично":
            for cell in row:
                cell.fill = red_fill

        elif level == "Предупреждение":
            for cell in row:
                cell.fill = yellow_fill
    risks = analyze_procurement_risks(items)
    # Лист 4 — Итоги по поставщикам
    ws_summary = wb.create_sheet("Итоги поставщиков")

    ws_summary.append([
        "Поставщик",
        "Всего позиций в КП",
        "Позиций выиграно",
        "Сумма выигранных позиций",
        "Рисков",
        "Рейтинг",
        "Комментарий"
    ])

    best_supplier = None
    best_wins = -1

    for supplier, stat in supplier_stats.items():
        wins = stat.get("wins", 0)

        if wins > best_wins:
            best_wins = wins
            best_supplier = supplier
        supplier_risks = len([
            risk for risk in risks
            if risk.get("supplier") == supplier
        ])

        if supplier_risks <= 1:
            rating = "A"
        elif supplier_risks <= 3:
            rating = "B"
        else:
            rating = "C"
        comment = "Лидер по количеству минимальных цен" if wins > 0 else "Нет выигранных позиций"

        ws_summary.append([
            supplier,
            stat.get("positions", 0),
            wins,
            stat.get("sum_wins", 0),
            supplier_risks,
            rating,
            comment
        ])

    style_sheet(ws_summary, {
        "A": 35,
        "B": 20,
        "C": 20,
        "D": 25,
        "E": 12,
        "F": 12,
        "G": 45
    })

    for row_idx in range(2, ws_summary.max_row + 1):
        if ws_summary.cell(row=row_idx, column=1).value == best_supplier:
            for col_idx in range(1, 8):
                ws_summary.cell(row=row_idx, column=col_idx).fill = green_fill
    # Лист 5 — Риски закупки
    ws_risks = wb.create_sheet("Риски закупки")

    ws_risks.append([
        "Поставщик",
        "Риск",
        "Уровень",
        "Поле",
        "Значение",
        "Комментарий"
    ])



    for risk in risks:
        ws_risks.append([
            risk.get("supplier", "не указано"),
            risk.get("risk", ""),
            risk.get("level", ""),
            risk.get("field", ""),
            risk.get("value", ""),
            risk.get("comment", "")
        ])

    style_sheet(ws_risks, {
        "A": 30,
        "B": 35,
        "C": 15,
        "D": 25,
        "E": 35,
        "F": 80
    })
        # Лист 6 — Проверка поставщиков
    ws_supplier_check = wb.create_sheet("Проверка поставщиков")

    ws_supplier_check.append([
        "Поставщик",
        "Производитель",
        "Страна",
        "НДС",
        "Гарантия",
        "Предоплата",
        "Рисков",
        "Оценка"
    ])

    for item in items:
        supplier = item.get("supplier", "не указано")
        supplier_risks = [
            risk for risk in risks
            if risk.get("supplier") == supplier
        ]

        payment_terms = str(item.get("payment_terms", "")).lower()
        prepayment = "Да" if "100" in payment_terms or "полная предоплата" in payment_terms else "Нет"

        risks_count = len(supplier_risks)

        if risks_count <= 1:
            supplier_score = "Надёжный"
        elif risks_count <= 3:
            supplier_score = "Требует проверки"
        else:
            supplier_score = "Высокий риск"

        ws_supplier_check.append([
            supplier,
            item.get("manufacturer", "не указано"),
            item.get("country", "не указано"),
            item.get("vat", "не указано"),
            item.get("warranty", "не указано"),
            prepayment,
            risks_count,
            supplier_score
        ])

    style_sheet(ws_supplier_check, {
        "A": 35,
        "B": 30,
        "C": 18,
        "D": 18,
        "E": 20,
        "F": 15,
        "G": 12,
        "H": 25
    })
    # Лист 7 — Заключение
    ws_conclusion = wb.create_sheet("Заключение")

    total_positions = len(position_data)

    ws_conclusion.append(["Параметр", "Значение"])
    ws_conclusion.append(["Всего КП", len(items)])
    ws_conclusion.append(["Всего уникальных позиций", total_positions])
    ws_conclusion.append(["Лидер по минимальным ценам", best_supplier or "не определён"])
    ws_conclusion.append(["Потенциальная экономия", total_saving])
    ws_conclusion.append(["Выявлено рисков закупки", len(risks)])

    ws_conclusion.append(["", ""])
    ws_conclusion.append(["Рекомендация", ""])

    if best_supplier:
        recommendation = (
            f"По результатам анализа коммерческих предложений лидер по количеству "
            f"минимальных цен — {best_supplier}. Рекомендуется дополнительно проверить "
            f"сроки поставки, условия оплаты, наличие товара, гарантию и включение доставки."
        )
    else:
        recommendation = (
            "Победитель не определён. Рекомендуется проверить корректность цен и наименований позиций."
        )

    ws_conclusion.append(["Текст заключения", recommendation])

    style_sheet(ws_conclusion, {
        "A": 30,
        "B": 90
    })

    ws_conclusion["B8"].fill = yellow_fill

    wb.save(output_path)
